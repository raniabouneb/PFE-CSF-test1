"""Parsing des fichiers Excel pour l'import de présences (sans accès DB)."""

from __future__ import annotations

import io
from datetime import date, datetime

import openpyxl


def normalize_status(raw: str) -> str | None:
    """Normalise une valeur de statut saisie dans le fichier."""
    if raw is None:
        return None
    t = (
        str(raw)
        .strip()
        .lower()
        .replace("é", "e")
        .replace("è", "e")
        .replace("ê", "e")
    )
    if t in ("present", "oui", "o", "1"):
        return "present"
    if t in ("absent", "non", "n", "0"):
        return "absent"
    if t in ("excuse", "excused"):
        return "excused"
    return None


def _cell_to_str(value: object | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    return s if s else None


def parse_excel_file(file_bytes: bytes, group_member_emails: set[str]) -> list[dict]:
    """Lit la 1ʳᵉ feuille, détecte les colonnes par libellé (insensible à la casse)."""
    bio = io.BytesIO(file_bytes)
    wb = openpyxl.load_workbook(bio, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValueError("Colonnes manquantes : date, email, statut requis") from None

        col_map: dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            if cell is None:
                continue
            key = str(cell).strip().lower()
            if key in ("date", "heure", "email", "statut"):
                col_map[key] = idx

        for req in ("date", "email", "statut"):
            if req not in col_map:
                raise ValueError("Colonnes manquantes : date, email, statut requis")

        out: list[dict] = []
        row_num = 1

        def get_cell(data_row: tuple, col_key: str) -> str | None:
            i = col_map[col_key]
            if i >= len(data_row):
                return None
            return _cell_to_str(data_row[i])

        for data_row in rows_iter:
            row_num += 1
            if not data_row:
                continue
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in data_row):
                continue

            date_s = get_cell(data_row, "date")
            heure_s = get_cell(data_row, "heure") if "heure" in col_map else None
            email_cell = get_cell(data_row, "email")
            statut_raw = get_cell(data_row, "statut")

            email = (email_cell or "").strip().lower()
            display_email = (email_cell or "").strip()

            if not email or "@" not in email:
                out.append(
                    {
                        "rowIndex": row_num,
                        "date": date_s,
                        "heure": heure_s,
                        "email": display_email or email,
                        "statut": None,
                        "resolution": "error",
                        "errorMessage": "Email invalide",
                    }
                )
                continue

            norm_status = normalize_status(statut_raw or "")
            if norm_status is None:
                out.append(
                    {
                        "rowIndex": row_num,
                        "date": date_s,
                        "heure": heure_s,
                        "email": email,
                        "statut": None,
                        "resolution": "error",
                        "errorMessage": f"Statut inconnu: {statut_raw!r}",
                    }
                )
                continue

            resolution: str = "found" if email in group_member_emails else "new"
            out.append(
                {
                    "rowIndex": row_num,
                    "date": date_s,
                    "heure": heure_s,
                    "email": email,
                    "statut": norm_status,
                    "resolution": resolution,
                    "errorMessage": None,
                }
            )
        return out
    finally:
        wb.close()
